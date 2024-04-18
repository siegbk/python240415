import  openpyxl  as  op  

#샘플로 제공된 파일을 오픈한다. 
wb = op.load_workbook("result2.xlsx") 
#마지막으로 활성화된 시트를 가져온다
ws = wb.active

#방법 1 : Sheet의 Cell 속성 사용하기
data1 = ws.cell(row=1, column=2).value

#방법 2 : 엑셀 인덱스(Range) 사용하기
data2 = ws["C1"].value

#위 결과 출력해보기
print("cell(1,2) : ", data1)
print('Range("C1"):', data2)

#범위를 지정한 경우 
rng = ws["A5:A12"] 

for  rng_data  in  rng: 
    for  cell_data  in  rng_data: 
        print(cell_data.value) 

